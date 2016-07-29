from openpyxl import load_workbook
from openpyxl import Workbook

def saveData(data, fileName, sheetName='Sheet'):
    keys = sorted(data)

    try:
        wb = load_workbook(fileName)
        ws = wb.create_sheet(sheetName)
    except:
        wb = Workbook()
        ws = wb.active
        ws.title = sheetName
          
    if 'time' in keys:
        timeCheck=True
        temp = data['time']
        ws.cell(column = 1, row = 1, value = 'time')
        for rowIndex in range(2, len(temp) + 2):
            ws.cell(column = 1, row = rowIndex, value = temp[rowIndex - 2])
        keys.remove('time')
    else:
        timeCheck=False

    for colIndex in range(0,  len(keys) ):
        temp = data[keys[colIndex]]
        ws.cell(column = colIndex+1+int(timeCheck), row = 1, value = keys[colIndex])
        for rowIndex in range(2, len(temp) + 2):
            ws.cell(column = colIndex+1+int(timeCheck), row = rowIndex, value = temp[rowIndex - 2])
            
    wb.save(fileName)
    
def readData(fileName,sheetName):
    wb = load_workbook(fileName)
    ws = wb.get_sheet_by_name(sheetName)
    data = {}

    for col in ws.columns:
        counter = 1
        key = ''
        colVals = []
        for cell in col:
            if counter == 1:
                key = cell.value
                data[key] = key
            else:
                colVals.append(cell.value)

            counter += 1
        data[key] = colVals
    return data

def getSheetNames(fileName):
    wb=load_workbook(fileName)
    SheetNames=wb.get_sheet_names()
    return SheetNames



# Test Run

if __name__=='__main__':
    import Plot
    data = {'time':[1,2,3,4], 'Dev1/ai0':[5,6,7,8], 'C':[9,10,11,12,13,14,15,15,64,2345],'D':[1,2,43,5,5,6,3,43,54,2,52,245,245,2542,45,524,4,524],'E':[3241,2431,41231423,1432,1234,4,526,413],'F':[1,2,43,5,6,7,8,8]}
    [x, y]=Plot.GenData(2, 100, 0, 1, 5)
    [x, y2] = Plot.GenData(3, 100, -1, 3, 5)
    #data = {'time': x, 'Chan1': y, 'Chan2': y2}
    #data={'Chan1':y,'Chan2':y2}
    #keys = sorted(data)

    saveData(fileName='PotatoTestData2.xlsx', sheetName='Sheet', data=data)
